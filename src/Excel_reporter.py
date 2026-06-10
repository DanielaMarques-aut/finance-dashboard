import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import date
import os
from src.formater import (style_header, style_curency_column, auto_adjust_column_width, add_alternating_row_colors,
freeze_header, add_spending_by_category_chart, COLORS)

def build_excel_report(df, summary, output_path):
    """Builds an Excel report with summary and transaction details."""
    # Summary sheet data
    summary_data = [
        ["MONTHLY FINANCE REPORT", ""],
        [f"Generated: {date.today().strftime('%B %d, %Y')}", ""],
        ["", ""],
        ["Metric", "Value (€)"],
        ["Total Income", summary["income"]],
        ["Total Expenses", summary["expenses"]],
        ["Net Savings", summary["net"]],
        ["Total Transactions", summary["transaction_count"]],
        ["", ""],
        ["Biggest Expense", summary["biggest_expense"]["description"]],
        ["Amount", summary["biggest_expense"]["amount"]],
    ]
    print("✓ Prepared summary data for Excel report.")
    #transactions sheet data
     # Transactions sheet
    df_export = df[["date", "description", "amount", "category"]].copy()
    df_export["date"] = df_export["date"].dt.strftime("%Y-%m-%d")
    df_export.columns = ["Date", "Description", "Amount (€)", "Category"]
    print("✓ Prepared transactions data for Excel report.")
    
    # Categories sheet
    categories_df = pd.DataFrame([
        {"Category": cat, "Amount (€)": round(amount, 2)}
        for cat, amount in summary["by_category"].items()
    ])
    print("✓ Prepared categories data for Excel report.")
    
    # --- Write to Excel ---
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    print("✓ Created reports directory.")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Summary sheet — manual data, not a DataFrame
        pd.DataFrame(summary_data).to_excel(
            writer, sheet_name="Summary",
            index=False, header=False
        )
        
        # Transactions sheet
        df_export.to_excel(
            writer, sheet_name="Transactions",
            index=False
        )
        
        # Categories sheet
        categories_df.to_excel(
            writer, sheet_name="Categories",
            index=False
        )
    print("✓ Written data to Excel file.")
    # --- Format with openpyxl ---
    wb = load_workbook(output_path)
    
    # Format Summary sheet
    ws_summary = wb["Summary"]
    ws_summary["A1"].font = Font(bold=True, size=16, color=COLORS["header_bg"])
    ws_summary["A4"].font = Font(bold=True, color="FFFFFF")
    ws_summary["B4"].font = Font(bold=True, color="FFFFFF")
    style_header(ws_summary, 4, 2)
    style_curency_column(ws_summary, "B", 5, 7)
    style_curency_column(ws_summary, "B", 11, 11)
    
    ws_summary["B7"].font = Font(
        bold=True,
        color="1D9E75" if summary["net"] > 0 else "DC2626"
    )
    auto_adjust_column_width(ws_summary, 2)
    print("✓ Formatted Summary sheet.")
    
    # Format Transactions sheet
    ws_trans = wb["Transactions"]
    style_header(ws_trans, 1, 4)
    freeze_header(ws_trans)
    style_curency_column(ws_trans, "C", 2, ws_trans.max_row)
    add_alternating_row_colors(ws_trans, 2, ws_trans.max_row, 4)
    print("✓ Styled Transactions sheet with headers, currency formatting, and alternating row colors.")
    # Color transactions: green for income, red for expenses
    for row in range(2, ws_trans.max_row + 1):
        amount_cell = ws_trans[f"C{row}"]
        if amount_cell.value and float(amount_cell.value) > 0:
            amount_cell.font = Font(color="166534", bold=True)
        elif amount_cell.value:
            amount_cell.font = Font(color="DC2626")
    
    auto_adjust_column_width(ws_trans, 4)
    print("✓ Formatted Transactions sheet.")
    
    # Format Categories sheet
    ws_cats = wb["Categories"]
    style_header(ws_cats, 1, 2)
    freeze_header(ws_cats)
    style_curency_column(ws_cats, "B", 2, ws_cats.max_row)
    add_alternating_row_colors(ws_cats, 2, ws_cats.max_row, 2)
    auto_adjust_column_width(ws_cats, 2)
    print("✓ Formatted Categories sheet.")
    # Add chart to Categories sheet
    wb = add_spending_by_category_chart("Categories", wb)
    print("✓ Added spending by category chart to Categories sheet.")
    
    # Set sheet order — Summary first
    wb.active = wb["Summary"]
    
    wb.save(output_path)
   
    print(f"✓ Excel report saved to {output_path}")
    
    return output_path