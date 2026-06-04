from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart

# Color palette
COLORS = {
    "header_bg":   "4F46E5",   # indigo
    "header_text": "FFFFFF",   # white
    "income_bg":   "DCFCE7",   # light green
    "expense_bg":  "FEE2E2",   # light red
    "alt_row":     "F8F7FF",   # very light purple
    "border":      "E2E0F9",   # light indigo
}
def style_header(worksheet, row_num, col_num):
    """Applies styling to the header row."""
    for col in range(1, col_num + 1):
        cell = worksheet.cell(row=row_num, column=col)
        cell.font = Font(bold=True, color=COLORS["header_text"], size=12)
        cell.fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(border_style="thin", color=COLORS["border"]))


def style_curency_column(worksheet, col_letter,start_row, end_row):
    """Formats a column as currency."""
    for row in range(start_row, end_row + 1):  # Skip header
        worksheet[f"{col_letter}{row}"].number_format = numbers.FORMAT_CURRENCY_EUR_SIMPLE


def auto_adjust_column_width(worksheet, col_num):
    """Auto-adjusts column widths based on content."""
    for col in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 30)  # Limit max width to 30 characters


def add_alternating_row_colors(worksheet, start_row, end_row, col_num):
    """Adds alternating row colors for better readability."""
    for row in range(start_row, end_row + 1):
        if row % 2 == 0:
    
            for col in range(1, col_num + 1):
                    worksheet.cell(row=row, column=col).fill = PatternFill(
                        start_color=COLORS["alt_row"],
                        end_color=COLORS["alt_row"],
                        fill_type="solid"
                    )
def freeze_header(worksheet):
    """Freezes the header row."""
    worksheet.freeze_panes = "A2"   

def add_spending_by_category_chart(category_sheet, workbook):
    """Adds a bar chart for spending by category."""
    worksheet= workbook[category_sheet]
    max_row = worksheet.max_row
    
      
    # Create the bar chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.width = 20
    chart.height = 12
    chart.title = "Spending by Category"
    chart.x_axis.title = "Category"
    chart.y_axis.title = "Amount (€)"
    
    data = Reference(worksheet, min_col=2, min_row=1, max_row=max_row)
    categories_ref = Reference(worksheet, min_col=1, min_row=2, max_row=max_row)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories_ref)
    
    # Position the chart
    worksheet.add_chart(chart, "D2")
    return workbook