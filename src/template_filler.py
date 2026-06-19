
from openpyxl import load_workbook
from datetime import date
import os

def fill_Template(summary, df, output_path, monthly_df):
    """Fills the Excel template with summary and transaction data."""
    # Load the template
   
    wb = load_workbook("data/Finance_template.xlsx")
    print("laded woerkbook")
    # Fill Dashboard sheet
    ws_dashboard = wb["Dashboard"]
    moth_label=date.today().strftime("%B %Y")
    ws_dashboard["B2"] = moth_label
    ws_dashboard["B3"] = date.today().strftime("%B %d, %Y")
    ws_dashboard["B7"] = round(summary["income"], 2)
    ws_dashboard["B8"] = round(summary["expenses"], 2)
    ws_dashboard["B9"] = round(summary["net"], 2)
    # Apply number formatting and simple styling
    try:
        from openpyxl.styles import numbers, Font, Alignment
        ws_dashboard["B7"].number_format = numbers.FORMAT_CURRENCY_EUR_SIMPLE
        ws_dashboard["B8"].number_format = numbers.FORMAT_CURRENCY_EUR_SIMPLE
        ws_dashboard["B9"].number_format = numbers.FORMAT_CURRENCY_EUR_SIMPLE
        ws_dashboard["B7"].font = Font(bold=True)
        ws_dashboard["B8"].font = Font(bold=True)
        # Net: green if positive, red if negative
        ws_dashboard["B9"].font = Font(bold=True, color=("1D9E75" if summary.get("net",0) > 0 else "DC2626"))
        ws_dashboard["B7"].alignment = Alignment(horizontal="right")
        ws_dashboard["B8"].alignment = Alignment(horizontal="right")
        ws_dashboard["B9"].alignment = Alignment(horizontal="right")
    except Exception:
        pass

    
    # Fill Transactions sheet
    ws_trans = wb["Transactions"]
    df_export = df[["date", "description", "amount", "category"]].copy()
    df_export["date"] = df_export["date"].dt.strftime("%Y-%m-%d")
    for r_idx, row in enumerate(df_export.values, start=2):
        ws_trans[f"A{r_idx}"] = row[0]
        ws_trans[f"B{r_idx}"] = row[1]
        ws_trans[f"C{r_idx}"] = row[2]
        ws_trans[f"D{r_idx}"] = row[3]

# Fill Categories sheet
    ws_cats = wb["Categories"]
    total_expenses = summary.get("expenses") or sum(summary.get("by_category", {}).values())
    for r_idx, (cat, amount) in enumerate(summary["by_category"].items(), start=2):
        ws_cats[f"A{r_idx}"] = cat
        ws_cats[f"B{r_idx}"] = round(amount, 2)
        # write percentage of total in column C
        try:
            perc = (amount / total_expenses) if total_expenses else 0
        except Exception:
            perc = 0
        ws_cats[f"C{r_idx}"] = perc
        try:
            from openpyxl.styles import numbers
            ws_cats[f"B{r_idx}"].number_format = numbers.FORMAT_CURRENCY_EUR_SIMPLE
            ws_cats[f"C{r_idx}"].number_format = '0.0%'
        except Exception:
            pass
        print("filled data")
#monthly comparation sheet

    if monthly_df is not None:
        #creates montly comparation sheet
        ws_monthly = wb.create_sheet("Monthly Comparison")
        print("Created sheet")
        #creates Headrs (include Net column)
        headers= ["Month", "Income €", "Expenses €", "Transactions", "Net €"]
        for col, header in enumerate(headers, 1):
            cell = ws_monthly.cell(row=1, column=col, value=header)
            cell.font = cell.font.copy(bold=True)
            from openpyxl.styles import PatternFill, Alignment
            cell.fill = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        print("Created headers")
        #Data to rows
        for row_idx, row in enumerate(monthly_df.values, start=2):
            for col_idx, col_value in enumerate(row,start=1):
                ws_monthly.cell(row=row_idx, column=col_idx, value=col_value)
        # Format columns: B and C as currency, D as integer
        from openpyxl.styles import numbers
        for r in range(2, ws_monthly.max_row + 1):
            try:
                ws_monthly[f"B{r}"].number_format = numbers.FORMAT_CURRENCY_EUR_SIMPLE
                ws_monthly[f"C{r}"].number_format = numbers.FORMAT_CURRENCY_EUR_SIMPLE
                ws_monthly[f"D{r}"].number_format = '0'
                # Net column E as currency
                ws_monthly[f"E{r}"].number_format = numbers.FORMAT_CURRENCY_EUR_SIMPLE
            except Exception:
                pass
        print("loaded data!")


#save the filled template as a new file
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"✓ Excel report generated at {output_path}")